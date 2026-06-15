from __future__ import annotations

import json
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


ROOT = Path("/Users/songbingyan/Desktop/只有免费商用字体26.6.13整理")
LINKS_FILE = Path("/Users/songbingyan/Desktop/字体库_官方链接补充_第一批_第二批合集 (1).xlsx")
OUT = Path(__file__).resolve().parents[1] / "src" / "data" / "fonts.json"

FONT_EXTS = {".ttf", ".otf", ".ttc", ".woff", ".woff2"}

FIGMA_SEEDS = [
    ("chinese", "handwriting", "马克笔手写", "Nishiki-teki"),
    ("chinese", "handwriting", "优设好身体", "YOUSHEhaoshenti"),
    ("chinese", "handwriting", "优设鲨鱼菲特健康体", "YouSheShaYuFeiTeJiankangTi"),
    ("chinese", "handwriting", "阿里健康体", "Alibaba Health Font 2.0 PY"),
    ("chinese", "handwriting", "阿里健康体", "Alibaba Health Font 2.0 CN"),
    ("chinese", "handwriting", "阿里健康体 2.0 盲文", "Alibaba Health Font2.0Braille"),
    ("chinese", "handwriting", "龚帆免费体", "gongfanmianfeiti2.0"),
    ("chinese", "handwriting", "鸿雷板书简体", "honglei sim"),
    ("chinese", "handwriting", "千图笔锋手写体", "qiantubifengshouxieti"),
    ("chinese", "handwriting", "沐瑶软笔手写体", "MuyaoPleased"),
    ("chinese", "handwriting", "平方上上谦体", "平方上上谦体"),
    ("chinese", "handwriting", "演示悠然小楷", "slideyouran"),
    ("chinese", "handwriting", "沐瑶随心手写体", "Muyao-Softbrush"),
    ("chinese", "calligraphy", "玉葱楷書激無料版", "玉葱楷書激無料版"),
    ("chinese", "calligraphy", "光良酒干杯体", "光良酒-干杯体"),
    ("chinese", "calligraphy", "演示斜黑体", "Source-KeynoteartHans"),
    ("chinese", "calligraphy", "杨任东竹石体", "YRDZST"),
    ("chinese", "calligraphy", "阿里妈妈刀隶体", "Alimama DaoLiTi"),
    ("chinese", "calligraphy", "庞门正道真贵楷体", "PangZhenGui-PMzD"),
    ("chinese", "calligraphy", "方正楷体简体", "FZKai-Z03S"),
    ("chinese", "calligraphy", "方正楷体GBK", "FZKai-Z03"),
    ("chinese", "calligraphy", "方正楷体繁体", "FZKai-Z03T"),
    ("chinese", "calligraphy", "演示夏行楷", "Slidexiaxing"),
    ("chinese", "calligraphy", "演示秋鸿楷", "Slideqiuhong"),
    ("chinese", "calligraphy", "演示春风楷", "Slidechunfeng"),
    ("chinese", "calligraphy", "演示佛系体", "Slidefu"),
    ("chinese", "calligraphy", "刻石錄颜體", "I.Ngaan"),
    ("chinese", "calligraphy", "阿里妈妈东方大楷", "Alimama DongFangDaKai"),
    ("chinese", "calligraphy", "字库江湖古风体", "字库江湖古风体"),
    ("chinese", "calligraphy", "华康楷体", "DFKaiGB Std"),
    ("chinese", "calligraphy", "华康隶书体", "DFLiShuGB Std"),
    ("chinese", "calligraphy", "华康正颜楷体", "DFYanKaiGB Std"),
    ("chinese", "calligraphy", "华康唐风隶", "DFTanLiGB Std"),
    ("chinese", "cute", "胖胖猪肉体", "07NikumaruFont"),
    ("chinese", "cute", "字魂扁桃体", "zihunbiantaoti"),
    ("chinese", "cute", "美呗嘿嘿体", "美呗嘿嘿体"),
    ("chinese", "cute", "白无常可可体", "BWCKKT"),
    ("chinese", "cute", "jf open 粉圓", "jf-openhuninn-2.0"),
    ("chinese", "cute", "优设字由棒棒体", "YS HelloFont BangBangTi"),
    ("chinese", "cute", "站酷快乐体", "HappyZcool-2016"),
    ("chinese", "cute", "仓耳小丸子", "TsangerxWz"),
]


CATEGORY_DIRS = {
    "01-常用字体": ("common", "all"),
    "02-中文字体": ("chinese", "all"),
    "03-英文字体": ("english", "all"),
    "04-数字字体": ("number", "all"),
    "05-多语言字体": ("multilingual", "all"),
}

MULTILINGUAL_DIRS = {
    "00-多语言全部": "all",
    "01-繁中": "traditional",
    "02-日文": "japanese",
    "03-韩文": "korean",
    "04-越南文": "vietnamese",
    "05-印尼语": "indonesian",
    "06-阿拉伯文": "arabic",
}


def compact(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def key(value: str) -> str:
    value = compact(value).lower()
    value = re.sub(r"[\s_\-./]+", "", value)
    value = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", value)
    return value


def slug(value: str) -> str:
    value = compact(value).lower()
    value = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "font"


def clean_stem(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"[-_](\d{5,}|copy|\d+)$", "", stem, flags=re.I)
    stem = re.sub(r"[_]+", " ", stem)
    stem = re.sub(r"\s+", " ", stem).strip()
    return stem


def display_name_from_stem(stem: str) -> str:
    name = stem
    name = re.sub(r"\.(ttf|otf|ttc|woff2?)$", "", name, flags=re.I)
    name = re.sub(r"[-_](Regular|Medium|Bold|Light|Heavy|Thin|Black|SemiBold|ExtraLight|ExtraBold|Italic)$", r" \1", name, flags=re.I)
    return compact(name)


def normalized_figma_name(name: str) -> str:
    name = compact(name)
    if not name or "?" in name or "？" in name or set(name) <= {"—", "-", " "}:
        return ""
    name = name.replace("_", " ")
    name = re.sub(r"\s+", " ", name)
    return name


def infer_chinese_subcategory(text: str) -> str:
    lower = text.lower()
    rules = [
        ("songti", ["宋", "serif", "明", "书宋", "仿宋", "sourcehanserif", "noto serif"]),
        ("handwriting", ["手写", "marker", "script", "hand", "马克", "笔", "芳华", "agile", "lingdong"]),
        ("calligraphy", ["楷", "隶", "行", "书法", "毛笔", "刀隶", "竹石", "颜", "真贵", "kai", "kaiti", "daoli", "dakai", "qingsong", "zhengui"]),
        ("cute", ["可爱", "可可", "丸子", "胖", "快乐", "棒棒", "嘿嘿", "cute", "xiaowanzi", "nikumaru"]),
        ("round", ["圆", "圓", "rounded", "yuan", "round"]),
        ("title", ["标题", "display", "logo", "heavy", "bold", "黑体", "金刚", "庞门", "优设", "站酷", "粗"]),
        ("heiti", ["黑", "sans", "hei", "gothic", "雅黑", "普惠", "鸿蒙", "harmony", "oppo", "misans"]),
    ]
    for sub, needles in rules:
        if any(n.lower() in lower for n in needles):
            return sub
    return "heiti"


def infer_english_subcategory(text: str) -> str:
    lower = text.lower()
    rules = [
        ("mono", ["mono", "code", "cascadia", "jetbrains", "consola", "courier"]),
        ("serif", ["serif", "slab", "merriweather", "lora", "arvo", "bitter", "georgia"]),
        ("rounded", ["round", "rounded", "baloo", "nunito", "quicksand"]),
        ("display", ["display", "black", "poster", "bangers", "bahiana", "monoton", "title", "headline", "sketch"]),
        ("ui", ["inter", "roboto", "sf pro", "segoe", "encode", "system", "ui"]),
        ("sans-serif", ["sans", "gothic", "open", "montserrat", "poppins", "lato", "work"]),
    ]
    for sub, needles in rules:
        if any(n in lower for n in needles):
            return sub
    return "sans-serif"


def preview_for(category: str) -> str:
    if category == "number":
        return "1234567890"
    if category == "english":
        return "Aa"
    if category == "multilingual":
        return "Global Type\n多语言样张"
    return "字有风骨\n设计有温度"


def read_link_rows() -> list[dict]:
    if not LINKS_FILE.exists():
        return []
    wb = load_workbook(LINKS_FILE, data_only=True, read_only=True)
    rows: list[dict] = []
    for sheet_name in ("合集_官方链接数据", "第一批_去重字体", "第二批_英文去重字体"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        try:
            headers = [compact(v) for v in next(iterator)]
        except StopIteration:
            continue
        for row in iterator:
            item = {headers[i]: compact(row[i]) if i < len(row) else "" for i in range(len(headers))}
            figma_name = item.get("Figma名称 figmaName", "")
            name = item.get("字体展示名 name", "")
            official_url = item.get("officialUrl", "")
            if not figma_name and not name:
                continue
            rows.append(
                {
                    "name": name or figma_name,
                    "figmaName": normalized_figma_name(figma_name or name),
                    "officialUrl": official_url,
                    "sourceType": item.get("sourceType", ""),
                    "trustLevel": item.get("可信度", ""),
                    "needVerify": item.get("needVerify", ""),
                    "note": item.get("备注", ""),
                    "rawCategory": item.get("一级分类", item.get("分类", "")),
                    "rawSubCategory": item.get("二级分类/标签", item.get("二级分类", "")),
                }
            )
    return rows


def infer_category_from_link(row: dict) -> tuple[str, str]:
    raw = row.get("rawCategory", "") + " " + row.get("rawSubCategory", "")
    text = raw + " " + row["name"] + " " + row["figmaName"]
    if "常用" in raw:
        return "common", "all"
    if "数字" in raw:
        return "number", "all"
    if "多语言" in raw:
        return "multilingual", "all"
    if "中文" in raw:
        return "chinese", infer_chinese_subcategory(text)
    if "英文" in raw:
        sub_map = {
            "ui": "ui",
            "无衬线": "sans-serif",
            "衬线": "serif",
            "圆角": "rounded",
            "等宽": "mono",
            "标题": "display",
        }
        for source, sub in sub_map.items():
            if source in raw.lower():
                return "english", sub
        return "english", infer_english_subcategory(text)
    return "english", infer_english_subcategory(text)


def file_category(path: Path) -> tuple[str, str]:
    rel = path.relative_to(ROOT)
    top = rel.parts[0]
    category, sub = CATEGORY_DIRS.get(top, ("english", "all"))
    if category == "multilingual" and len(rel.parts) > 2:
        sub = MULTILINGUAL_DIRS.get(rel.parts[1], "all")
    stem = clean_stem(path)
    if category == "chinese":
        sub = infer_chinese_subcategory(stem)
    elif category == "english":
        sub = infer_english_subcategory(stem)
    return category, sub


def add_or_merge(store: dict[str, dict], item: dict) -> None:
    dedupe_key = key(item["figmaName"]) or key(item["name"])
    existing = store.get(dedupe_key)
    if existing:
        for field in ("officialUrl", "sourceType", "trustLevel", "needVerify", "note", "fontFile"):
            if not existing.get(field) and item.get(field):
                existing[field] = item[field]
        existing["tags"] = sorted(set(existing.get("tags", [])) | set(item.get("tags", [])))
        route_key = f'{item["category"]}:{item["subCategory"]}'
        if route_key not in existing["routes"]:
            existing["routes"].append(route_key)
        if (
            existing["category"] not in {"common", "number"}
            and existing["category"] == item["category"]
            and existing["subCategory"] == "all"
            and item["subCategory"] != "all"
        ):
            existing["subCategory"] = item["subCategory"]
        return
    item["id"] = slug(item["figmaName"])
    item["routes"] = [f'{item["category"]}:{item["subCategory"]}']
    store[dedupe_key] = item


def main() -> None:
    links = read_link_rows()
    by_key: dict[str, dict] = {}
    for row in links:
        for value in (row["figmaName"], row["name"]):
            by_key[key(value)] = row

    store: dict[str, dict] = {}

    for path in sorted(ROOT.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in FONT_EXTS:
            continue
        stem = clean_stem(path)
        match = by_key.get(key(stem)) or by_key.get(key(display_name_from_stem(stem)))
        category, sub = file_category(path)
        name = match["name"] if match else display_name_from_stem(stem)
        figma_name = match["figmaName"] if match else normalized_figma_name(stem)
        item = {
            "id": "",
            "name": name,
            "figmaName": figma_name,
            "category": category,
            "subCategory": sub,
            "previewText": preview_for(category),
            "officialUrl": match.get("officialUrl", "") if match else "",
            "sourceType": match.get("sourceType", "") if match else "",
            "trustLevel": match.get("trustLevel", "") if match else "",
            "needVerify": match.get("needVerify", "") if match else "",
            "note": match.get("note", "") if match else "",
            "fontFile": str(path),
            "tags": sorted(set([category, sub, name, figma_name, "免费商用"])),
        }
        add_or_merge(store, item)

    for row in links:
        category, sub = infer_category_from_link(row)
        item = {
            "id": "",
            "name": row["name"],
            "figmaName": row["figmaName"],
            "category": category,
            "subCategory": sub,
            "previewText": preview_for(category),
            "officialUrl": row["officialUrl"],
            "sourceType": row["sourceType"],
            "trustLevel": row["trustLevel"],
            "needVerify": row["needVerify"],
            "note": row["note"],
            "fontFile": "",
            "tags": sorted(set([category, sub, row["name"], row["figmaName"], "免费商用"])),
        }
        add_or_merge(store, item)

    for category, sub, name, figma_name in FIGMA_SEEDS:
        figma_name = normalized_figma_name(figma_name) or name
        match = by_key.get(key(figma_name)) or by_key.get(key(name)) or {}
        item = {
            "id": "",
            "name": name,
            "figmaName": figma_name,
            "category": category,
            "subCategory": sub,
            "previewText": preview_for(category),
            "officialUrl": match.get("officialUrl", ""),
            "sourceType": match.get("sourceType", ""),
            "trustLevel": match.get("trustLevel", ""),
            "needVerify": match.get("needVerify", ""),
            "note": match.get("note", ""),
            "fontFile": "",
            "tags": sorted(set([category, sub, name, figma_name, "免费商用"])),
        }
        add_or_merge(store, item)

    fonts = sorted(store.values(), key=lambda f: (f["category"], f["subCategory"], f["name"].lower()))
    used_ids: dict[str, int] = {}
    for font in fonts:
        base = font["id"]
        count = used_ids.get(base, 0)
        used_ids[base] = count + 1
        if count:
            font["id"] = f"{base}-{count + 1}"
        font["encodedFigmaName"] = quote(font["figmaName"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(fonts, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {len(fonts)} fonts to {OUT}")


if __name__ == "__main__":
    main()
